from io import BytesIO
from typing import List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel

from ..core.deps import get_current_user
from ..db.session import get_db

router = APIRouter(prefix="/api/projects", tags=["projects"])


class ProjectCreate(BaseModel):
    project_name: str
    wbs_number: Optional[str] = None
    solution_product: Optional[str] = None
    company: str
    location: str
    status: str = "active"
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    assignee_ids: Optional[List[int]] = None


class ImportSummary(BaseModel):
    projects_created: int
    projects_updated: int
    milestones_created: int
    issues_created: int
    progress_created: int
    warnings: List[str]


def _require_admin(current_user):
    if not current_user["is_admin"]:
        raise HTTPException(403, "Admin only.")


def _normalize_text(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_csv(value) -> list[str]:
    text = _normalize_text(value)
    if not text:
        return []
    return [part.strip() for part in text.split(",") if part.strip()]


def _sheet_rows(sheet):
    header_row = next(sheet.iter_rows(values_only=True), None)
    if not header_row:
        return
    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
    for row_index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(value is None or str(value).strip() == "" for value in values):
            continue
        yield row_index, {
            headers[idx]: values[idx] if idx < len(values) else None
            for idx in range(len(headers))
        }


def _set_example_sheet_style(sheet):
    sheet.freeze_panes = "A2"
    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells) + 2
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width, 14), 34)


def _build_import_template() -> BytesIO:
    workbook = Workbook()

    projects_ws = workbook.active
    projects_ws.title = "Projects"
    projects_ws.append([
        "project_key", "project_name", "wbs_number", "solution_product", "company",
        "location", "status", "start_date", "end_date", "assignee_emails",
    ])
    projects_ws.append([
        "PRJ-ALPHA", "Alpha MES Upgrade", "WBS-24001", "MES", "Yokogawa",
        "Seoul", "active", "2026-04-01", "2026-09-30", "admin@example.com, pm@example.com",
    ])

    milestones_ws = workbook.create_sheet("Milestones")
    milestones_ws.append(["project_key", "title", "planned_date", "actual_date", "status"])
    milestones_ws.append(["PRJ-ALPHA", "Kickoff completed", "2026-04-05", "2026-04-05", "done"])
    milestones_ws.append(["PRJ-ALPHA", "Factory UAT", "2026-06-15", "", "planned"])

    issues_ws = workbook.create_sheet("Issues")
    issues_ws.append([
        "project_key", "issue_key", "title", "status", "priority",
        "start_date", "end_date", "details",
    ])
    issues_ws.append([
        "PRJ-ALPHA", "ISSUE-NETWORK", "PLC network latency", "진행중", "high",
        "2026-04-07", "", "Intermittent latency observed during peak traffic.",
    ])

    progress_ws = workbook.create_sheet("IssueProgress")
    progress_ws.append(["project_key", "issue_key", "title", "start_date", "end_date", "details"])
    progress_ws.append([
        "PRJ-ALPHA", "ISSUE-NETWORK", "Packet capture completed", "2026-04-08", "2026-04-08",
        "Captured switch-to-PLC traffic and isolated two spike windows.",
    ])

    instructions_ws = workbook.create_sheet("Instructions")
    instructions_ws.append(["Sheet", "How to use"])
    instructions_ws.append(["Projects", "One row per project. project_key links all sheets. Existing projects match by WBS first, then project name."])
    instructions_ws.append(["Milestones", "Optional. Add project milestones linked by project_key."])
    instructions_ws.append(["Issues", "Optional. issue_key is required if you will import issue progress rows."])
    instructions_ws.append(["IssueProgress", "Optional. Each row must match both project_key and issue_key from the Issues sheet."])
    instructions_ws.append(["Dates", "Use YYYY-MM-DD format."])
    instructions_ws.append(["Project status", "Use active, on_hold, completed, or cancelled."])
    instructions_ws.append(["Milestone status", "Use planned, done, delayed, or cancelled."])

    for sheet in workbook.worksheets:
        _set_example_sheet_style(sheet)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _with_assignees(conn, projects: list) -> list:
    result = []
    for p in projects:
        assignees = conn.execute(
            """SELECT u.id, u.name, r.name as rank_name
               FROM project_assignments pa
               JOIN users u ON u.id=pa.user_id
               JOIN ranks r ON r.id=u.rank_id
               WHERE pa.project_id=? AND u.is_deleted=0
               ORDER BY u.name""",
            (p["id"],),
        ).fetchall()
        result.append({**p, "assignees": assignees})
    return result


def _find_user_ids(conn, emails: list[str], row_index: int, warnings: list[str]) -> list[int]:
    user_ids: list[int] = []
    for email in emails:
        user = conn.execute(
            "SELECT id FROM users WHERE lower(email)=lower(?) AND is_deleted=0",
            (email,),
        ).fetchone()
        if not user:
            warnings.append(f"Projects row {row_index}: assignee email '{email}' was not found and was skipped.")
            continue
        user_ids.append(user["id"])
    return user_ids


def _resolve_existing_project(conn, project_key: str, project_name: Optional[str], wbs_number: Optional[str]):
    if wbs_number:
        row = conn.execute(
            "SELECT id FROM projects WHERE wbs_number=? AND is_deleted=0",
            (wbs_number,),
        ).fetchone()
        if row:
            return row["id"]
    if project_name:
        row = conn.execute(
            "SELECT id FROM projects WHERE project_name=? AND is_deleted=0",
            (project_name,),
        ).fetchone()
        if row:
            return row["id"]
    row = conn.execute(
        "SELECT id FROM projects WHERE wbs_number=? AND is_deleted=0",
        (project_key,),
    ).fetchone()
    if row:
        return row["id"]
    row = conn.execute(
        "SELECT id FROM projects WHERE project_name=? AND is_deleted=0",
        (project_key,),
    ).fetchone()
    if row:
        return row["id"]
    return None


def _import_workbook(conn, workbook, current_user) -> ImportSummary:
    required_sheets = {"Projects", "Milestones", "Issues", "IssueProgress"}
    missing = [name for name in required_sheets if name not in workbook.sheetnames]
    if missing:
        raise HTTPException(400, f"Missing sheet(s): {', '.join(missing)}")

    warnings: list[str] = []
    project_ids_by_key: dict[str, int] = {}
    issue_ids_by_key: dict[tuple[str, str], int] = {}

    projects_created = 0
    projects_updated = 0
    milestones_created = 0
    issues_created = 0
    progress_created = 0

    for row_index, row in _sheet_rows(workbook["Projects"]):
        project_key = _normalize_text(row.get("project_key"))
        project_name = _normalize_text(row.get("project_name"))
        company = _normalize_text(row.get("company"))
        location = _normalize_text(row.get("location"))
        if not project_key or not project_name or not company or not location:
            raise HTTPException(400, f"Projects row {row_index}: project_key, project_name, company, and location are required.")

        body = ProjectCreate(
            project_name=project_name,
            wbs_number=_normalize_text(row.get("wbs_number")),
            solution_product=_normalize_text(row.get("solution_product")),
            company=company,
            location=location,
            status=_normalize_text(row.get("status")) or "active",
            start_date=_normalize_text(row.get("start_date")),
            end_date=_normalize_text(row.get("end_date")),
            assignee_ids=_find_user_ids(conn, _parse_csv(row.get("assignee_emails")), row_index, warnings),
        )

        existing_id = _resolve_existing_project(conn, project_key, body.project_name, body.wbs_number)
        if existing_id:
            conn.execute(
                """UPDATE projects
                   SET project_name=?, wbs_number=?, solution_product=?, company=?, location=?,
                       status=?, start_date=?, end_date=?, updated_by=?
                   WHERE id=? AND is_deleted=0""",
                (
                    body.project_name,
                    body.wbs_number,
                    body.solution_product,
                    body.company,
                    body.location,
                    body.status,
                    body.start_date,
                    body.end_date,
                    current_user["id"],
                    existing_id,
                ),
            )
            conn.execute("DELETE FROM project_assignments WHERE project_id=?", (existing_id,))
            for assignee_id in body.assignee_ids or []:
                conn.execute(
                    "INSERT OR IGNORE INTO project_assignments(project_id,user_id) VALUES(?,?)",
                    (existing_id, assignee_id),
                )
            project_id = existing_id
            projects_updated += 1
        else:
            conn.execute(
                """INSERT INTO projects
                       (project_name,wbs_number,solution_product,company,location,
                        status,start_date,end_date,created_by,updated_by)
                   VALUES(?,?,?,?,?,?,?,?,?,?)""",
                (
                    body.project_name,
                    body.wbs_number,
                    body.solution_product,
                    body.company,
                    body.location,
                    body.status,
                    body.start_date,
                    body.end_date,
                    current_user["id"],
                    current_user["id"],
                ),
            )
            project_id = conn.execute("SELECT last_insert_rowid() as id").fetchone()["id"]
            for assignee_id in body.assignee_ids or []:
                conn.execute(
                    "INSERT OR IGNORE INTO project_assignments(project_id,user_id) VALUES(?,?)",
                    (project_id, assignee_id),
                )
            projects_created += 1

        project_ids_by_key[project_key] = project_id

    for row_index, row in _sheet_rows(workbook["Milestones"]):
        project_key = _normalize_text(row.get("project_key"))
        title = _normalize_text(row.get("title"))
        planned_date = _normalize_text(row.get("planned_date"))
        if not project_key or not title or not planned_date:
            raise HTTPException(400, f"Milestones row {row_index}: project_key, title, and planned_date are required.")
        project_id = project_ids_by_key.get(project_key)
        if not project_id:
            raise HTTPException(400, f"Milestones row {row_index}: unknown project_key '{project_key}'.")
        conn.execute(
            """INSERT INTO project_milestones
                   (project_id, title, planned_date, actual_date, status, created_by, updated_by)
               VALUES (?,?,?,?,?,?,?)""",
            (
                project_id,
                title,
                planned_date,
                _normalize_text(row.get("actual_date")),
                _normalize_text(row.get("status")) or "planned",
                current_user["id"],
                current_user["id"],
            ),
        )
        milestones_created += 1

    for row_index, row in _sheet_rows(workbook["Issues"]):
        project_key = _normalize_text(row.get("project_key"))
        issue_key = _normalize_text(row.get("issue_key"))
        title = _normalize_text(row.get("title"))
        start_date = _normalize_text(row.get("start_date"))
        if not project_key or not issue_key or not title or not start_date:
            raise HTTPException(400, f"Issues row {row_index}: project_key, issue_key, title, and start_date are required.")
        project_id = project_ids_by_key.get(project_key)
        if not project_id:
            raise HTTPException(400, f"Issues row {row_index}: unknown project_key '{project_key}'.")
        conn.execute(
            """INSERT INTO project_issues
                   (project_id, title, status, priority, start_date, end_date, details, created_by, updated_by)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (
                project_id,
                title,
                _normalize_text(row.get("status")) or "초안",
                _normalize_text(row.get("priority")) or "normal",
                start_date,
                _normalize_text(row.get("end_date")),
                _normalize_text(row.get("details")),
                current_user["id"],
                current_user["id"],
            ),
        )
        issue_id = conn.execute("SELECT last_insert_rowid() as id").fetchone()["id"]
        issue_ids_by_key[(project_key, issue_key)] = issue_id
        issues_created += 1

    for row_index, row in _sheet_rows(workbook["IssueProgress"]):
        project_key = _normalize_text(row.get("project_key"))
        issue_key = _normalize_text(row.get("issue_key"))
        title = _normalize_text(row.get("title"))
        start_date = _normalize_text(row.get("start_date"))
        if not project_key or not issue_key or not title or not start_date:
            raise HTTPException(400, f"IssueProgress row {row_index}: project_key, issue_key, title, and start_date are required.")
        issue_id = issue_ids_by_key.get((project_key, issue_key))
        if not issue_id:
            raise HTTPException(400, f"IssueProgress row {row_index}: no imported issue matched project_key '{project_key}' and issue_key '{issue_key}'.")
        conn.execute(
            """INSERT INTO project_issue_progress
                   (issue_id, title, start_date, end_date, details, created_by, updated_by)
               VALUES (?,?,?,?,?,?,?)""",
            (
                issue_id,
                title,
                start_date,
                _normalize_text(row.get("end_date")),
                _normalize_text(row.get("details")),
                current_user["id"],
                current_user["id"],
            ),
        )
        progress_created += 1

    return ImportSummary(
        projects_created=projects_created,
        projects_updated=projects_updated,
        milestones_created=milestones_created,
        issues_created=issues_created,
        progress_created=progress_created,
        warnings=warnings,
    )


@router.get("")
def list_projects(
    status: Optional[str] = None,
    q: Optional[str] = None,
    mine: Optional[bool] = None,
    current_user=Depends(get_current_user),
):
    query = """SELECT p.*, d.name as dept_name
               FROM projects p
               LEFT JOIN departments d ON d.id=p.department_id
               WHERE p.is_deleted=0"""
    params: list = []
    if mine:
        query += """ AND EXISTS (
            SELECT 1 FROM project_assignments pa
            WHERE pa.project_id=p.id AND pa.user_id=?)"""
        params.append(current_user["id"])
    if status:
        query += " AND p.status=?"
        params.append(status)
    if q:
        query += " AND p.project_name LIKE ?"
        params.append(f"%{q}%")
    query += " ORDER BY p.project_name"

    with get_db() as conn:
        projects = conn.execute(query, params).fetchall()
        return _with_assignees(conn, projects)


@router.post("")
def create_project(body: ProjectCreate, current_user=Depends(get_current_user)):
    with get_db() as conn:
        conn.execute(
            """INSERT INTO projects
                   (project_name,wbs_number,solution_product,company,location,
                    status,start_date,end_date,created_by,updated_by)
               VALUES(?,?,?,?,?,?,?,?,?,?)""",
            (
                body.project_name,
                body.wbs_number,
                body.solution_product,
                body.company,
                body.location,
                body.status,
                body.start_date,
                body.end_date,
                current_user["id"],
                current_user["id"],
            ),
        )
        pid = conn.execute("SELECT last_insert_rowid() as id").fetchone()["id"]
        if body.assignee_ids:
            for uid in body.assignee_ids:
                try:
                    conn.execute(
                        "INSERT INTO project_assignments(project_id,user_id) VALUES(?,?)",
                        (pid, uid),
                    )
                except Exception:
                    pass
        return conn.execute("SELECT * FROM projects WHERE id=?", (pid,)).fetchone()


@router.put("/{project_id}")
def update_project(
    project_id: int,
    body: ProjectCreate,
    current_user=Depends(get_current_user),
):
    with get_db() as conn:
        conn.execute(
            """UPDATE projects
               SET project_name=?,wbs_number=?,solution_product=?,
                   company=?,location=?,status=?,start_date=?,end_date=?,
                   updated_by=?
               WHERE id=? AND is_deleted=0""",
            (
                body.project_name,
                body.wbs_number,
                body.solution_product,
                body.company,
                body.location,
                body.status,
                body.start_date,
                body.end_date,
                current_user["id"],
                project_id,
            ),
        )
        if body.assignee_ids is not None:
            conn.execute(
                "DELETE FROM project_assignments WHERE project_id=?",
                (project_id,),
            )
            for uid in body.assignee_ids:
                try:
                    conn.execute(
                        "INSERT INTO project_assignments(project_id,user_id) VALUES(?,?)",
                        (project_id, uid),
                    )
                except Exception:
                    pass
        return conn.execute("SELECT * FROM projects WHERE id=?", (project_id,)).fetchone()


@router.get("/import/template")
def download_project_import_template(current_user=Depends(get_current_user)):
    _require_admin(current_user)
    output = _build_import_template()
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="project-import-template.xlsx"'},
    )


@router.post("/import")
async def import_projects_excel(
    file: UploadFile = File(...),
    current_user=Depends(get_current_user),
):
    _require_admin(current_user)
    if not file.filename:
        raise HTTPException(400, "Please upload an Excel file.")

    try:
        workbook = load_workbook(BytesIO(await file.read()))
    except Exception as exc:
        raise HTTPException(400, "The uploaded file could not be read as an Excel workbook.") from exc

    with get_db() as conn:
        return _import_workbook(conn, workbook, current_user)
